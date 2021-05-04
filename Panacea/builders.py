import datetime
import itertools

from django import forms
from django.db import transaction
from django.db.models import Sum
from django.forms import modelformset_factory, BaseModelFormSet, ModelForm
from django.http import Http404
from django.shortcuts import redirect
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment

from Panacea.models import service_offered, revenue, transit_data, expense, fund_balance, revenue_source, \
    transit_metrics, expense_source, fund_balance_type, transit_mode, summary_organization_progress, organization
from Panacea.utilities import get_current_summary_report_year, get_all_data_steps_completed


def None_sum(*args):
    args = [a for a in args if not a is None]
    return sum(args) if args else 0


class SummaryBuilder:
    '''Basic building block for other classes it holds REPORT_TYPES'''

    def __init__(self):
        self.REPORT_TYPES = ['transit_data', 'revenue', 'expense', 'fund_balance']
        self.REVENUE_TYPES = ['Operating', 'Capital', 'Other']
        self.REVENUE_GOV = ['Local', 'State', 'Federal', 'Other']


class SummaryBuilderReportType(SummaryBuilder):
    '''This class contains key building instructions for all builders based on the summary builder REPORT_TYPE'''

    def __init__(self, report_type):
        super().__init__()
        self.report_type = report_type

    def get_model(self):
        '''returns the appropriate model for the given report type'''
        if self.report_type == "revenue":
            return revenue
        elif self.report_type == "transit_data":
            return transit_data
        elif self.report_type == "expense":
            return expense
        elif self.report_type == "fund_balance":
            return fund_balance
        elif self.report_type == 'organization':
            return organization
        else:
            raise Http404("Report type does not exist. -4")

    def get_model_data(self, filter_inactive=True):
        '''returns the appropriate model objects for the given report type'''
        if self.report_type == "revenue":
            return revenue.objects.exclude(revenue_source__inactive_flag=filter_inactive)
        elif self.report_type == "transit_data":
            return transit_data.objects
        elif self.report_type == "expense":
            return expense.objects
        elif self.report_type == "fund_balance":
            return fund_balance.objects
        elif self.report_type == 'organization':
            return organization.objects
        else:
            raise Http404("Report type does not exist. -4")

    def get_metric_model(self):
        '''returns the appropriate metric model for the given report type'''
        if self.report_type == "revenue":
            return revenue_source
        elif self.report_type == "transit_data":
            return transit_metrics
        elif self.report_type == "expense":
            return expense_source
        elif self.report_type == "fund_balance":
            return fund_balance_type
        elif self.report_type == 'organization':
            return organization
        else:
            raise Http404("Report type does not exist. -5")

    def get_metric_model_data(self, filter_inactive=True):
        '''returns the appropriate metric model objects for the given report type'''
        if self.report_type == "revenue":
            return revenue_source.objects.exclude(inactive_flag=filter_inactive)
        elif self.report_type == "transit_data":
            return transit_metrics.objects
        elif self.report_type == "expense":
            return expense_source.objects
        elif self.report_type == "fund_balance":
            return fund_balance_type.objects
        elif self.report_type == 'organization':
            return organization.objects
        else:
            raise Http404("Report type does not exist. -5")

    def get_metric_model_name(self):
        '''Returns the metric model as a string. Useful for certain queries and other operations'''
        if self.report_type == "revenue":
            return 'revenue_source'
        elif self.report_type == "transit_data":
            return 'transit_metric'
        elif self.report_type == "expense":
            return 'expense_source'
        elif self.report_type == "fund_balance":
            return 'fund_balance_type'
        else:
            raise Http404("Report type does not exist. -6")

    def get_metric_id_field_name(self):
        '''Returns the name of the field name of the id field in the metric model as a string.'''
        if self.report_type == "transit_data":
            return 'transit_metric_id'
        else:
            metric_model = self.get_metric_model()
            return metric_model.__name__ + '_id'

    def get_pretty_report_type(self):
        if self.report_type == "revenue":
            return 'Revenue'
        elif self.report_type == "transit_data":
            return 'Transit Data'
        elif self.report_type == "expense":
            return 'Expenses'
        elif self.report_type == "fund_balance":
            return 'Fund balances'
        elif self.report_type == 'organization':
            return "Organization"
        else:
            raise Http404("Report type does not exist. -6")

    def has_order_in_summary(self):
        if self.report_type in ['transit_data', 'revenue', 'expense']:
            return True
        else:
            return False

    def has_headings_in_summary(self):
        if self.report_type in ["expense"]:
            return True
        else:
            return False


class SummaryDataEntryBuilder(SummaryBuilderReportType):
    '''This class constructs all of the forms needed to collect summary data.

    For some REPORT_TYPEs the data collection form are split into multiple pages these are filtered using either one or
    two parameters depending on the REPORT_TYPE'''

    def __init__(self, report_type, target_organization, form_filter_1=None, form_filter_2=None):
        super().__init__(report_type)

        # self.report_type = report_type  # reports can be about revenue, transit data, expenses, and ending fund balances
        self.target_organization = target_organization  # the org submitting a report
        self.year = get_current_summary_report_year()  # TODO this function needs to be updated
        self.form_filter_1 = form_filter_1  # Forms can be filtered by the selectors at the top of the page for example reporting based on direct operated, fixed route transit
        self.form_filter_2 = form_filter_2
        # These control how the form moves to the next form
        self.max_form_increment = 0  # if the max increment is meet it will move to the next report type, otherwise it will go to the next set of filters
        self.current_increment = 0

        self.nav_filter_count, self.nav_filters = self.get_header_navigation()
        self.set_default_form_filters()  # sets the starting filters for the form
        self.set_max_form_increment()
        self.set_current_increment()

    def set_default_form_filters(self):
        if self.form_filter_1 is not None:
            pass
        else:
            # TODO create ordering for metric types
            if self.report_type == "revenue":
                self.form_filter_1 = self.nav_filters[0][0]
                self.form_filter_2 = self.nav_filters[0][1]
            elif self.report_type == "transit_data":
                # TODO Make this into something that makes more sense
                self.form_filter_1 = service_offered.objects.filter(organization=self.target_organization).order_by(
                    'transit_mode_id').values_list('transit_mode__name').first()[0]
                self.form_filter_2 = service_offered.objects.filter(organization=self.target_organization).order_by(
                    'transit_mode_id').values_list('administration_of_mode').first()[0]
                # self.form_filter_1 = self.get_model().objects.filter(organization=self.target_organization).order_by(
                #     'transit_mode__name').values_list('transit_mode__name').first()[0]
                # self.form_filter_2 = self.get_model().objects.filter(organization=self.target_organization).order_by(
                #     'administration_of_mode').values_list('administration_of_mode').first()[0]
            elif self.report_type in ["expense", "fund_balance"]:
                self.form_filter_1 = None
                self.form_filter_2 = None
            else:
                raise Http404("Report type does not exist. -1")

    def set_max_form_increment(self):
        '''returns the appropriate model for the given report type'''
        # if self.report_type == "revenue":
        #     self.max_form_increment = 7
        # elif self.report_type == "transit_data":
        #     self.max_form_increment = service_offered.objects.filter(organization=self.target_organization).count()
        # elif self.report_type == "expense":
        #     self.max_form_increment = 1
        # elif self.report_type == "fund_balance":
        #     self.max_form_increment = 1
        # else:
        #     raise Http404("Report type does not exist. -2")

        if isinstance(self.nav_filters, list):
            self.max_form_increment = len(self.nav_filters)
        else:
            self.max_form_increment = 1

    def set_current_increment(self):
        '''returns the appropriate model for the given report type'''
        if self.report_type == "revenue":
            self.current_increment = self.nav_filters.index([self.form_filter_1, self.form_filter_2]) + 1
        elif self.report_type == "transit_data":
            self.current_increment = self.nav_filters.index([self.form_filter_1, self.form_filter_2]) + 1
        elif self.report_type == "expense":
            self.current_increment = 1
        elif self.report_type == "fund_balance":
            self.current_increment = 1
        else:
            raise Http404("Report type does not exist. -3")

    def get_all_metric_ids(self):
        '''Returns a distinct list of all metric ids that are needed given the agency classification, if applicable.'''
        classification = self.target_organization.summary_organization_classifications
        if self.report_type in ['transit_data', 'revenue', ]:
            metric_ids = list(
                self.get_metric_model_data().filter(agency_classification=classification).values_list('id',
                                                                                                      flat=True).distinct())
        elif self.report_type in ['fund_balance', 'expense', ]:
            metric_ids = list(self.get_metric_model_data().values_list('id', flat=True).distinct())
        else:
            raise Http404
        return metric_ids

    def get_create_metric_dictionary(self, metric):
        '''Used to create a new empty instance of a report metric. So an empty form may be displayed'''
        create_dictionary = {}
        if self.report_type in ['transit_data', ]:
            create_dictionary = {'year': metric[1],
                                 'organization': self.target_organization,
                                 'transit_mode': transit_mode.objects.get(name=self.form_filter_1),
                                 'administration_of_mode': self.form_filter_2,
                                 self.get_metric_id_field_name(): metric[0],
                                 'reported_value': None,
                                 }
        elif self.report_type in ['revenue', 'expense', 'fund_balance', ]:
            create_dictionary = {'year': metric[1],
                                 'organization': self.target_organization,
                                 self.get_metric_id_field_name(): metric[0],
                                 'reported_value': None,
                                 }
        else:
            raise Http404
        return create_dictionary

    def get_or_create_all_form_metrics(self):
        '''Gets all reported form metrics applicable to the form type, organization, and year.  If the metric has not been reported it creates it.'''
        model = self.get_model_data()
        if self.report_type == "transit_data":
            report_model = model.filter(transit_mode__name=self.form_filter_1,
                                        administration_of_mode=self.form_filter_2)
        else:
            report_model = model
        field_id = self.get_metric_id_field_name()

        current_report_metric_ids = list(report_model.filter(organization=self.target_organization,
                                                             year__gte=self.year - 2).values_list(field_id,
                                                                                                  'year').distinct())
        all_report_metric_ids = self.get_all_metric_ids()
        all_metric_ids_and_years = list(
            itertools.product(all_report_metric_ids, [self.year, self.year - 1, self.year - 2]))
        if len(current_report_metric_ids) != len(all_metric_ids_and_years):

            all_metric_ids_and_years = set(map(tuple, all_metric_ids_and_years))
            current_report_metric_ids = set(map(tuple, current_report_metric_ids))
            missing_metrics = list(all_metric_ids_and_years - current_report_metric_ids)
            # missing_metrics = all_metric_ids_and_years.symmetric_difference(current_report_metric_ids)
            # TODO there are some metrics that are currently filtered out that orgs previously reported on. How do we want to deal with these?
            if len(missing_metrics) > 0:
                with transaction.atomic():
                    for m in missing_metrics:
                        model.create(**self.get_create_metric_dictionary(m))

        form_metrics = model.filter(organization=self.target_organization).order_by(
            self.get_metric_id_field_name() + '__name')
        return form_metrics

    def get_widgets(self):
        '''Used to build widgets dynamically based on form type.'''

        if self.report_type == 'transit_data':
            widget_attrs = {'class': 'form-control validate-field', 'autocomplete': "off"}
        else:
            widget_attrs = {'class': 'form-control grand-total-sum validate-field', 'onchange': 'findTotal_wrapper();',
                            'autocomplete': "off"}

        widgets = {'id': forms.NumberInput(),
                   self.get_metric_model_name(): forms.Select(),
                   'year': forms.NumberInput(),
                   'reported_value': forms.TextInput(attrs=widget_attrs),
                   'comments': forms.Textarea(attrs={'class': 'form-control comment-field', "rows": 3})
                   }
        return widgets

    def create_model_formset_factory(self):
        '''Creates a fromset factory based on the information contained in the class information'''
        my_formset_factory = modelformset_factory(self.get_model(), form=ModelForm, formfield_callback=None,
                                                  formset=BaseModelFormSet, extra=0, can_delete=False,
                                                  can_order=False, max_num=None,
                                                  fields=["id", self.get_metric_model_name(), "year", "reported_value",
                                                          "comments"],
                                                  exclude=None,
                                                  widgets=self.get_widgets(),
                                                  validate_max=False, localized_fields=None,
                                                  labels=None, help_texts=None, error_messages=None,
                                                  min_num=None, validate_min=False, field_classes=None)

        return my_formset_factory

    def get_formset_query_dict(self):
        '''Builds a dynamic dictionary used for querying the aproriate metrics giving the filter criteria and organization classification'''
        if self.report_type in ['transit_data', ]:

            query_dict = {'transit_mode__name': self.form_filter_1,
                          'administration_of_mode': self.form_filter_2,
                          'transit_metric__agency_classification': self.target_organization.summary_organization_classifications
                          }
        elif self.report_type in ['revenue', ]:
            query_dict = {'revenue_source__government_type': self.form_filter_1,
                          'revenue_source__funding_type': self.form_filter_2,
                          'revenue_source__agency_classification': self.target_organization.summary_organization_classifications}
        elif self.report_type in ['expense', ]:
            query_dict = {
                'expense_source__agency_classification': self.target_organization.summary_organization_classifications}
        elif self.report_type in ['fund_balance', ]:
            query_dict = {
                'fund_balance_type__agency_classification': self.target_organization.summary_organization_classifications}
        else:
            raise Http404
        return query_dict

    def get_form_queryset(self):
        form_querysets = self.get_or_create_all_form_metrics()
        form_querysets = form_querysets.filter(**self.get_formset_query_dict())
        return form_querysets

    def get_formsets_labels_masking_class_and_help_text(self):
        '''Builds formsets by year with labels and masking classes'''
        my_formset_factory = self.create_model_formset_factory()
        form_querysets = self.get_form_queryset()

        formsets = {}
        i = 0
        for year_x in ['this_year', 'previous_year', 'two_years_ago']:
            formsets[year_x] = my_formset_factory(
                queryset=form_querysets.filter(year=self.year - i),
                prefix=year_x)
            i += 1
        formset_labels = form_querysets.filter(year=self.year).values_list(
            self.get_metric_model_name() + "__name", flat=True)
        help_text = form_querysets.filter(year=self.year).values_list(
            self.get_metric_model_name() + "__help_text", flat=True)
        if self.report_type != "transit_data":
            masking_class = ['Money'] * len(formset_labels)
        else:
            masking_class = form_querysets.filter(year=self.year).values_list(
                self.get_metric_model_name() + "__form_masking_class", flat=True)

        return formsets, formset_labels, masking_class, help_text

    def get_other_measure_totals(self):
        '''Gets totals from that need to be aggrigated on the page but are not presented due to the filters on the form.'''
        if self.report_type == 'transit_data':
            return None

        total_not_this_form = {}
        if self.get_formset_query_dict() == {}:
            for year_x in ['this_year', 'previous_year', 'two_years_ago']:
                total_not_this_form[year_x] = {'reported_value__sum': 0}
        else:
            report_model_data = self.get_model_data()
            total_not_this_form_queryset = report_model_data.filter(organization=self.target_organization).exclude(
                **self.get_formset_query_dict())
            i = 0
            for year_x in ['this_year', 'previous_year', 'two_years_ago']:
                total_not_this_form[year_x] = total_not_this_form_queryset.filter(year=self.year - i).aggregate(
                    Sum('reported_value'))
                if total_not_this_form[year_x]['reported_value__sum'] == None:
                    total_not_this_form[year_x]['reported_value__sum'] = 0
                i += 1

        return total_not_this_form

    def get_header_navigation(self):
        '''gets the data needed to build header navigation for filters'''
        if self.report_type in ['transit_data', ]:
            filter_count = 2
            my_services_offered = service_offered.objects.filter(organization=self.target_organization).order_by(
                'transit_mode_id')
            filters = []
            for service in my_services_offered:
                filters.append([service.transit_mode.name, service.administration_of_mode])
        elif self.report_type in ['revenue', ]:
            filter_count = 2
            revenues = revenue_source.objects.filter(
                agency_classification=self.target_organization.summary_organization_classifications).values(
                'government_type', 'funding_type').distinct()

            filters = []
            for source in revenues:
                filters.append([source['government_type'], source['funding_type']])

            if len(filters) > 1:
                revenue_list_order_type = self.REVENUE_TYPES
                revenue_list_order_gov = self.REVENUE_GOV
                filters.sort(key=lambda x: revenue_list_order_type.index(x[1]))
                filters.sort(key=lambda x: revenue_list_order_gov.index(x[0]))


        elif self.report_type in ['expense', 'fund_balance', ]:
            filter_count = 0
            if self.report_type == 'expense':
                filters = 'Expenses'
            elif self.report_type == 'fund_balance':
                filters = "Ending fund balances"
        else:
            raise Http404
        return filter_count, filters

    def save_with_post_data(self, post_data):
        my_formset_factory = self.create_model_formset_factory()
        query_sets = self.get_form_queryset()
        i = 0
        for year_x in ['this_year', 'previous_year', 'two_years_ago']:
            query = query_sets.filter(year=self.year - i).order_by(
                self.get_metric_id_field_name())
            formset = my_formset_factory(post_data, queryset=query_sets.filter(year=self.year - i).order_by(
                self.get_metric_id_field_name()), prefix=year_x)
            for form in formset:
                if form.is_valid():
                    form.save()
                else:
                    print(form.errors)

            i += 1

    def go_to_next_form(self, save_only=False):
        if save_only:
            return redirect('summary_reporting_filters', self.report_type, self.form_filter_1, self.form_filter_2)
        next_increment = self.current_increment + 1
        if next_increment > self.max_form_increment:
            org_progress = summary_organization_progress.objects.get(organization=self.target_organization)
            if self.report_type == 'fund_balance':
                org_progress.ending_balances = True
                org_progress.save()
                if get_all_data_steps_completed(self.target_organization.id):
                    return redirect('submit_data')
                else:
                    return redirect('you_skipped_a_step')
            else:
                if self.report_type == "revenue":
                    org_progress.revenue = True
                    org_progress.save()
                elif self.report_type == "transit_data":
                    org_progress.transit_data = True
                    org_progress.save()
                elif self.report_type == "expense":
                    org_progress.expenses = True
                    org_progress.save()
                elif self.report_type == "fund_balance":
                    raise Http404("Report type does not exist. -7a")
                else:
                    raise Http404("Report type does not exist. -7b")

                new_report_type = self.REPORT_TYPES[self.REPORT_TYPES.index(self.report_type) + 1]
                return redirect('summary_reporting_type', new_report_type)
        else:
            self.form_filter_1 = self.nav_filters[self.current_increment][0]
            self.form_filter_2 = self.nav_filters[self.current_increment][1]
            return redirect('summary_reporting_filters', self.report_type, self.form_filter_1, self.form_filter_2)

    def get_next_builder(self):
        self.current_increment = self.current_increment + 1
        if self.current_increment > self.max_form_increment:
            try:
                new_report_type = self.REPORT_TYPES[self.REPORT_TYPES.index(self.report_type) + 1]
                new_form_filter_1 = None
                new_form_filter_2 = None
            except:
                return False
        else:
            new_report_type = self.report_type
            new_form_filter_1 = self.nav_filters[self.current_increment - 1][0]
            new_form_filter_2 = self.nav_filters[self.current_increment - 1][1]

        new_builder = SummaryDataEntryBuilder(new_report_type, self.target_organization, new_form_filter_1,
                                              new_form_filter_2)

        return new_builder

class SummaryDataEntryTemplateData:
    '''Class that uses the SummaryDataEntryBuilder to create data needed in the template.'''

    def __init__(self, summary_data_entry_builder):
        self.formsets, self.formset_labels, self.masking_class, self.help_text = summary_data_entry_builder.get_formsets_labels_masking_class_and_help_text()
        self.report_type = summary_data_entry_builder.report_type
        self.year = summary_data_entry_builder.year
        self.form_filter_1 = summary_data_entry_builder.form_filter_1
        self.form_filter_2 = summary_data_entry_builder.form_filter_2
        self.other_totals = summary_data_entry_builder.get_other_measure_totals()
        self.masking_types = []
        self.nav_filter_count = summary_data_entry_builder.nav_filter_count
        self.nav_filters = summary_data_entry_builder.nav_filters

        if summary_data_entry_builder.report_type == "transit_data":
            self.show_totals = False
        else:
            self.show_totals = True

        if len(self.formset_labels) == 0:
            self.no_reports = True
        else:
            self.no_reports = False


class ConfigurationBuilder(SummaryBuilderReportType):
    '''Builder that constructs the data reporting configuration forms based on REPORT_TYPE.

    NOTE: Configuration forms have an additional REPORT_TYPE: organization'''

    def __init__(self, report_type, help_text=None):
        super().__init__(report_type)
        self.REPORT_TYPES = ['organization', 'transit_data', 'revenue', 'expense', 'fund_balance']
        self.help_text = help_text

        if self.help_text is None:
            self.primary_field_name, self.other_fields_list = self.get_model_fields()
        else:
            self.primary_field_name = "help_text"
            self.other_fields_list = []

    def get_model_data(self):
        super(ConfigurationBuilder, self).get_metric_model_data(filter_inactive=False)

    def get_metric_model_data(self):
        super(ConfigurationBuilder, self).get_metric_model_data(filter_inactive=False)

    def get_model_fields(self):
        '''returns the appropriate model for the given report type'''
        if self.report_type == "revenue":
            primary_field_name = 'agency_classification'
            other_fields_list = ['funding_type', 'government_type', 'inactive_flag']
            return primary_field_name, other_fields_list
        elif self.report_type == "transit_data":
            primary_field_name = 'agency_classification'
            other_fields_list = []
            return primary_field_name, other_fields_list
        # TODO add in expense and fund balance
        elif self.report_type == "expense":
            primary_field_name = 'agency_classification'
            other_fields_list = []
            return primary_field_name, other_fields_list
        elif self.report_type == "fund_balance":
            primary_field_name = 'agency_classification'
            other_fields_list = []
            return primary_field_name, other_fields_list
        elif self.report_type == "organization":
            primary_field_name = 'summary_organization_classifications'
            other_fields_list = []
            return primary_field_name, other_fields_list
        else:
            raise Http404("Report type does not exist. -4a")

    def get_data_relationship_one_2_one(self):
        if self.report_type in ['organization']:
            return True
        elif self.report_type in ['revenue', 'transit_data', 'expense', 'fund_balance']:
            return False
        else:
            return Http404("Report type does not exist. -4b")

    def create_model_formset_factory(self):
        '''Creates a fromset factory based on the information contained in the class information'''
        my_formset_factory = modelformset_factory(self.get_metric_model(), form=ModelForm, formfield_callback=None,
                                                  formset=BaseModelFormSet, extra=0, can_delete=False,
                                                  can_order=False, max_num=None,
                                                  fields=['id', "name",
                                                          self.primary_field_name] + self.other_fields_list,
                                                  exclude=None,
                                                  widgets=self.get_widgets(),
                                                  validate_max=False, localized_fields=None,
                                                  labels=None, help_texts=None, error_messages=None,
                                                  min_num=None, validate_min=False, field_classes=None)

        return my_formset_factory

    def get_query_set(self):
        print(self.report_type)

        query_set = self.get_metric_model().objects.order_by('name').all()
        return query_set

    def get_widgets(self):
        '''widgets'''
        if self.help_text is not None:
            widgets = {'id': forms.NumberInput(),
                       'name': forms.TextInput(attrs={'class': 'form-control AJAX_instant_submit',
                                                      'data-form-name': "summary_configure"}),
                       self.primary_field_name: forms.Textarea(attrs={'class': 'form-control AJAX_instant_submit',
                                                                      'data-form-name': "summary_configure",
                                                                      'rows': 2})}
            return widgets

        if self.get_data_relationship_one_2_one():
            widgets = {'id': forms.NumberInput(),
                       'name': forms.TextInput(attrs={'class': 'form-control AJAX_instant_submit',
                                                      'data-form-name': "summary_configure"}),
                       self.primary_field_name: forms.Select(attrs={'class': 'form-control AJAX_instant_submit',
                                                                    'data-form-name': "summary_configure"})}

        else:
            widgets = {'name': forms.TextInput(attrs={'class': 'form-control'}),
                       self.primary_field_name: forms.CheckboxSelectMultiple(
                           attrs={'class': 'form-check-inline no-bullet AJAX_instant_submit',
                                  'data-form-name': "summary_configure"})}

        if len(self.other_fields_list) > 0:
            for field in self.other_fields_list:
                widgets[field] = forms.Select(attrs={'class': 'form-control AJAX_instant_submit',
                                                     'data-form-name': "summary_configure"})

        return widgets

    def get_form_set(self, **kwargs):
        my_formset_factory = self.create_model_formset_factory()
        my_formset = my_formset_factory(queryset=self.get_query_set(), **kwargs)
        return my_formset


class ExportReport(SummaryBuilder):
    """Builder builds multiple export reports based on organization and has methods to output these reports in various
    formats"""

    def __init__(self, target_organization):
        super().__init__()
        self.target_organization = target_organization
        self.report_type_sub_report_dictionary = {}
        self.gather_report_data()

    def gather_report_data(self):
        """
        Gathers data for all report types.
        Data is stored in a dictionary, report_type_sub_report_dictionary, for each report type which holds a
        ExportReport_ReportType object.
        """

        current_builder = SummaryDataEntryBuilder(report_type=self.REPORT_TYPES[0],
                                                  target_organization=self.target_organization)
        while current_builder:
            data = ExportReport_Data(current_builder)
            if current_builder.report_type not in self.report_type_sub_report_dictionary:
                new_report_type = ExportReport_ReportType(current_builder.report_type)
                new_report_type.append_export_report_data(data)
                self.report_type_sub_report_dictionary[current_builder.report_type] = new_report_type
            else:
                self.report_type_sub_report_dictionary[current_builder.report_type].append_export_report_data(data)
            current_builder = current_builder.get_next_builder()

    def generate_excel_data_entry_report(self, file_save_os=False):
        year = get_current_summary_report_year()
        wb = Workbook()
        ws = wb.active
        bd = Side(style='medium', color="000000")
        ws.title = "SummaryReportData"
        ws.column_dimensions['A'].width = 47
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 10

        ws.append([self.target_organization.name])
        style_cell = ws.cell(row=ws.max_row, column=1)
        style_cell.font = Font(size=20, bold=True)
        ws.append(["Summary of Public Transportation Report data export run: {}".format(datetime.datetime.now())])
        style_cell = ws.cell(row=ws.max_row, column=1)
        style_cell.font = Font(italic=True)
        ws.append([])
        for report_type in self.report_type_sub_report_dictionary.values():

            for data in report_type.export_report_data_list:
                ws.append(["{} - {}".format(report_type.pretty_report_type, data.nav_headers)])
                style_cell = ws.cell(row=ws.max_row, column=1)
                style_cell.font = Font(bold=True)
                style_cell.border = Border(bottom=bd)
                ws.cell(row=ws.max_row, column=2).border = Border(bottom=bd)
                ws.cell(row=ws.max_row, column=3).border = Border(bottom=bd)
                ws.cell(row=ws.max_row, column=4).border = Border(bottom=bd)

                ws.append(['', year - 2, year - 1, year])
                for i in data.get_data(include_comment=False):
                    ws.append(i)
                ws.append([])
        if file_save_os:
            wb.save(filename="text.xlsx")
        else:
            return wb

    def summary_final_excel_report(self, file_save_os=False):
        pass

    def generate_annual_operating_table(self, include_comment=False):
        output = []

        for mode in self.report_type_sub_report_dictionary["transit_data"].export_report_data_list:
            # print(mode.nav_headers)
            output.append(mode.pretty_nav_headers)
            output.extend(mode.get_data(order_by_summary=True, remove_empty_data=True, include_comment=include_comment))

        return output

    def get_fare_revenue(self, include_comment=False):
        output = []
        temp_vanpool = False
        temp_farebox = ["Farebox revenues", 0, 0, 0]
        for mode in self.report_type_sub_report_dictionary["transit_data"].export_report_data_list:
            data = mode.get_data(include_comment=include_comment)
            # print(mode.nav_headers)
            if "Vanpool" in mode.nav_headers:
                # print('true')
                for i in data:
                    # print(i)
                    # print(i[1:])
                    if i[0] == "Farebox Revenues":
                        temp_vanpool = ['Vanpooling revenue', i[1], i[2], i[3]]
                        # print('temp_vanpool: ' + str(temp_vanpool))
            else:
                # print(data)
                for i in data:
                    # print(i)
                    # print(i[0])
                    if i[0] == "Farebox Revenues":
                        temp_farebox[1] = temp_farebox[1] + i[1]
                        temp_farebox[2] = temp_farebox[2] + i[2]
                        temp_farebox[3] = temp_farebox[3] + i[3]
        output.append(temp_farebox)
        if temp_vanpool:
            output.append(temp_vanpool)

        # print('output: ' + str(output))
        return [output]

    def generate_financial_information_table(self, include_comment=False):

        if self.target_organization.summary_organization_classifications.name in ["Transit", "Tribe", "Monorail", "Ferry"]:
            financial_table_output = {'Farebox revenue': self.get_fare_revenue(),
                                      'Operating revenue': [],
                                      'Other operating': [],
                                      'Other operating - subtotal': [],
                                      'Federal capital grant revenues': [],
                                      'State capital grant revenues': [],
                                      'Local capital expenditures': [],
                                      'Other expenditures': [],
                                      'Debt service': [],
                                      'Ending balances, December 31': []
                                      }
        else:
            raise NotImplementedError

        for value in financial_table_output['Farebox revenue']:
            financial_table_output['Operating revenue'].extend(value)
        del financial_table_output['Farebox revenue']

        for source in self.report_type_sub_report_dictionary['revenue'].export_report_data_list:
            if 'Operating' in source.nav_headers and 'Other' not in source.nav_headers:
                data = source.get_data(order_by_summary=True, remove_empty_data=True, include_comment=include_comment)
                financial_table_output['Operating revenue'].extend(data)

                # operating_revenue_subtotal = ['Other operating sub-total']
                # print('none: ' + str(financial_table_output['Operating revenue']))
            elif 'Operating' in source.nav_headers and 'Other' in source.nav_headers:
                data = source.get_data(order_by_summary=True, remove_empty_data=True, include_comment=include_comment)
                subtotal = ['Other operating sub-total']
                subtotal.extend(source.get_totals())

                if sum(subtotal[1:4]) > 0:
                    financial_table_output['Other operating - subtotal'].append(subtotal)
                    financial_table_output['Other operating'].extend(data)
                else:
                    del financial_table_output['Other operating']
                    del financial_table_output['Other operating - subtotal']
            elif 'Federal' in source.nav_headers and 'Capital' in source.nav_headers:
                data = source.get_data(order_by_summary=True, remove_empty_data=True, include_comment=include_comment)
                subtotal = ['Total federal capital']
                subtotal.extend(source.get_totals())

                if sum(subtotal[1:4]) > 0:
                    financial_table_output['Federal capital grant revenues'].extend(data)
                    financial_table_output['Federal capital grant revenues'].append(subtotal)
                else:
                    del financial_table_output['Federal capital grant revenues']
            elif 'State' in source.nav_headers and 'Capital' in source.nav_headers:
                data = source.get_data(order_by_summary=True, remove_empty_data=True, include_comment=include_comment)
                subtotal = ['Total state capital']
                subtotal.extend(source.get_totals())

                if sum(subtotal[1:4]) > 0:
                    financial_table_output['State capital grant revenues'].extend(data)
                    financial_table_output['State capital grant revenues'].append(subtotal)
                else:
                    del financial_table_output['State capital grant revenues']
            elif 'Local' in source.nav_headers and 'Capital' in source.nav_headers:
                data = source.get_data(order_by_summary=True, remove_empty_data=True, include_comment=include_comment)
                subtotal = ['Total local capital']
                subtotal.extend(source.get_totals())

                if sum(subtotal[1:4]) > 0:
                    financial_table_output['Local capital grant revenues'].extend(data)
                    financial_table_output['Local capital grant revenues'].append(subtotal)
                else:
                    del financial_table_output['Local capital grant revenues']
            elif 'Other' in source.nav_headers:
                data = source.get_data(order_by_summary=True, remove_empty_data=True, include_comment=include_comment)
                subtotal = ['Other Capital total']
                subtotal.extend(source.get_totals())

                if sum(subtotal[1:4]) > 0:
                    financial_table_output['Other capital'].extend(data)
                    financial_table_output['Other capital'].append(subtotal)
                else:
                    del financial_table_output['Other capital']

        # print('target dict: ' + str(financial_table_output['State capital grant revenues']))
        for source in self.report_type_sub_report_dictionary['expense'].export_report_data_list:
            data = source.get_data(order_by_summary=True, remove_empty_data=True, include_comment=include_comment)
            # print('data: ' + str(data))
            for row in data:
                # print('Row: ' + str(row))
                heading = expense_source.objects.get(name=row[0]).heading
                # print('heading: ' + str(heading))
                if heading == "Local capital expenditures":
                    financial_table_output['Local capital expenditures'].append(row)
                elif heading == "Other expenditures":
                    financial_table_output['Other expenditures'].append(row)
                elif heading == "Debt service":
                    financial_table_output['Debt service'].append(row)
            # print('dict: ' + str(financial_table_output['Local capital expenditures']))


        for source in self.report_type_sub_report_dictionary['fund_balance'].export_report_data_list:
            data = source.get_data(order_by_summary=True, remove_empty_data=True, include_comment=include_comment)
            subtotal = ['Ending balance total']
            subtotal.extend(source.get_totals())

            if sum(subtotal[1:4]) > 0:
                financial_table_output['Ending balances, December 31'].extend(data)
                financial_table_output['Ending balances, December 31'].append(subtotal)
            else:
                del financial_table_output['Ending balances, December 31']


        for key, value in {'Operating revenue': 'Total (excludes capital revenues)',
                           'Local capital expenditures': 'Total local capital',
                           'Debt service': 'Total debt service'}.items():

            total_row = self._calculate_total(value, financial_table_output[key])
            if total_row:
                if key =='Operating revenue':
                    if 'Other operating - subtotal' in financial_table_output.keys():
                        # print(financial_table_output['Other operating - subtotal'])
                        financial_table_output['Operating revenue'].extend(financial_table_output['Other operating - subtotal'])
                        total_row[1] = None_sum(total_row[1],
                                                financial_table_output['Other operating - subtotal'][0][1])
                        total_row[2] = None_sum(total_row[2],
                                                financial_table_output['Other operating - subtotal'][0][2])
                        total_row[3] = None_sum(total_row[3],
                                                financial_table_output['Other operating - subtotal'][0][3])
                        del financial_table_output['Other operating - subtotal']
                    if 'Other operating' in financial_table_output.keys():
                        financial_table_output['Operating revenue'].extend(financial_table_output['Other operating'])
                        del financial_table_output['Other operating']
                    financial_table_output['Operating revenue'].append(total_row)
                else:
                    if not total_row == ['']:
                        financial_table_output[key].append(total_row)

        if len(financial_table_output['Local capital expenditures']) == 0:
            del financial_table_output['Local capital expenditures']
        if len(financial_table_output['Other expenditures']) == 0:
            del financial_table_output['Other expenditures']
        if len(financial_table_output['Debt service']) == 0:
            del financial_table_output['Debt service']

        output = []

        for key in financial_table_output.keys():
            output.append(key)
            for i in financial_table_output[key]:
                # print(i)
                depth = lambda L: isinstance(L, list) and max(map(depth, L)) + 1
                if i is not None:
                    d = depth(i)
                    if d > 1:
                        output.extend(i)
                    else:
                        output.append(i)

        return output

    def _calculate_total(self, total_label, dict_value):
        total_row = [total_label, 0, 0, 0]
        # print(len(dict_value))
        # print(dict_value)
        for row in dict_value:
            # print(row)
            # print(len(row))
            if row is None:
                pass
            elif len(row) == 1:
                row = row[0]

            if "Depreciation" in row[0]:
                pass
            else:
                total_row[1] = None_sum(total_row[1], row[1])
                total_row[2] = None_sum(total_row[2], row[2])
                total_row[3] = None_sum(total_row[3], row[3])
        if total_row[1] == 0 and total_row[2] == 0 and total_row[3] == 0:
            return ['']
        else:
            return total_row

    def _calculate_percent(self, list_of_lists):
        # print(list_of_lists)
        for row in list_of_lists:
            if isinstance(row, list) and len(row) > 1:
                if row[2] == 0:
                    if row[3] == 0:
                        value_to_append = 0
                    else:
                        value_to_append = 100.00
                else:
                    value_to_append = (row[3]/row[2] - 1) * 100
                value_to_append = str("{:.2f}".format(round(value_to_append, 2)))
                if len(row) == 5:
                    row.append(row[4])
                    row[4] = value_to_append
                else:
                    row.append(value_to_append)
        return list_of_lists

    def _format_numbers(self, list_of_lists):
        non_money_transit_metrics = transit_metrics.objects.exclude(form_masking_class="Money").values_list('name', flat=True)

        for row in list_of_lists:
            # print(row)
            if isinstance(row, list) and len(row) > 1:
                if row[0] not in non_money_transit_metrics:
                    row[1] = "${:,.0f}".format(row[1])
                    row[2] = "${:,.0f}".format(row[2])
                    row[3] = "${:,.0f}".format(row[3])
                else:
                    if "FTE" in row[0]:
                        row[1] = "{:,.1f}".format(row[1])
                        row[2] = "{:,.1f}".format(row[2])
                        row[3] = "{:,.1f}".format(row[3])
                    else:
                        row[1] = "{:,.0f}".format(row[1])
                        row[2] = "{:,.0f}".format(row[2])
                        row[3] = "{:,.0f}".format(row[3])
        return list_of_lists

    def _format_report(self, list_of_lists):
        list_of_lists = self._calculate_percent(list_of_lists)
        list_of_lists = self._format_numbers(list_of_lists)
        return list_of_lists

    def generate_excel_summary_report(self, file_save_os=False, file_save_path='./', include_comment=False):
        year = get_current_summary_report_year()
        wb = Workbook()
        ws = wb.active
        bd = Side(style='thin', color="000000")
        ws.title = "SummaryReportData"
        ws.column_dimensions['A'].width = 47
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 10

        operating = self.generate_annual_operating_table(include_comment=include_comment)
        operating = self._format_report(operating)
        financial = self.generate_financial_information_table(include_comment=include_comment)

        financial = self._format_report(financial)

        totals_by_fund_source = self.generate_total_funds_by_source_table()

        ws.append([self.target_organization.name])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)



        ws.append(['Annual Operating Information',
                   get_current_summary_report_year()-2,
                   get_current_summary_report_year()-1,
                   get_current_summary_report_year()-0,
                   "One year change (%)"])
        for i in range(1, 6):
            ws.cell(row=ws.max_row, column=i).font = Font(bold=True)
            ws.cell(row=ws.max_row, column=i).alignment = Alignment(horizontal='center')
            ws.cell(row=ws.max_row, column=i).border = Border(bottom=bd, top=bd, right=bd, left=bd)

        first_item = True
        for i in operating:
            if isinstance(i, list):
                ws.append(i)
            else:
                if not first_item:
                    ws.append([""])
                else:
                    first_item=False
                ws.append([i])
                ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

        ws.append([''])
        ws.append(['Financial Information',
                   get_current_summary_report_year() - 2,
                   get_current_summary_report_year() - 1,
                   get_current_summary_report_year() - 0,
                   "One year change (%)"])
        for i in range(1, 6):
            ws.cell(row=ws.max_row, column=i).font = Font(bold=True)
            ws.cell(row=ws.max_row, column=i).alignment = Alignment(horizontal='center')

        first_item = True
        skip_bold=False
        for i in financial:
            if isinstance(i, list):
                ws.append(i)
                if "Depreciation" in i[0]:
                    skip_bold = True
            else:
                if not first_item or not skip_bold:
                    ws.cell(row=ws.max_row, column=1).alignment = Alignment(horizontal='right')
                else:
                    first_item = False
                if not skip_bold:
                    for j in range(1, 6):
                        ws.cell(row=ws.max_row, column=j).font = Font(bold=True)
                else:
                    skip_bold = False

                ws.append([i])
                ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
            if i[0] == "Other operating sub-total":
                ws.cell(row=ws.max_row, column=1).alignment = Alignment(horizontal='right')
                for j in range(1, 6):
                    ws.cell(row=ws.max_row, column=j).font = Font(bold=True)
            if "Other-" in i[0]:
                ws.cell(row=ws.max_row, column=1).alignment = Alignment(horizontal='right')
        ws.cell(row=ws.max_row, column=1).alignment = Alignment(horizontal='right')
        for i in range(1, 6):
            ws.cell(row=ws.max_row, column=i).font = Font(bold=True)

        ws.append([''])
        ws.append(['Total funds by source',
                   get_current_summary_report_year() - 2,
                   get_current_summary_report_year() - 1,
                   get_current_summary_report_year() - 0,
                   'One year change(%)'])
        for i in range(1, 6):
            ws.cell(row=ws.max_row, column=i).font = Font(bold=True)
            ws.cell(row=ws.max_row, column=i).alignment = Alignment(horizontal='center')

        # print(totals_by_fund_source)
        first_item = True
        for i in totals_by_fund_source:

            if isinstance(i, list) and len(i) > 1:
                ws.append(i)
            else:
                if not first_item:
                    ws.cell(row=ws.max_row, column=1).alignment = Alignment(horizontal='right')
                else:
                    first_item=False
                for j in range(1, 6):
                    ws.cell(row=ws.max_row, column=j).font = Font(bold=True)

                ws.append(i)
                ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        for i in range(1, 6):
            ws.cell(row=ws.max_row, column=i).font = Font(bold=True)
        ws.cell(row=ws.max_row, column=1).alignment = Alignment(horizontal='right')

        if file_save_os:
            save_name = file_save_path + self.target_organization.name + '.xlsx'
            wb.save(filename=save_name)
        else:
            return wb

    def generate_total_funds_by_source_table(self):
        i = 0
        output_dict = {'Local revenues': {},
                       'State revenues': {},
                       'Federal revenues': {},
                       'Operating investments': {},
                       'Local capital investments': {},
                       'State capital investments': {},
                       'Federal capital investments': {},
                       'Other investments': {}
                       }

        for year_x in ['this_year', 'previous_year', 'two_years_ago']:
            output_dict['Local revenues'][year_x] = revenue.objects.filter(
                organization=self.target_organization,
                year=get_current_summary_report_year() - i,
                revenue_source__government_type__in=["Local", "Other"]).aggregate(sum=Sum('reported_value'))['sum']
            output_dict['Local revenues'][year_x] = None_sum(output_dict['Local revenues'][year_x],
                                                             transit_data.objects.filter(
                                                                 organization=self.target_organization,
                                                                 year=get_current_summary_report_year() - i,
                                                                 transit_metric_id=10).aggregate(sum=Sum('reported_value'))['sum'])
            print('Local revenues')
            print(output_dict['Local revenues'][year_x])

            output_dict['State revenues'][year_x] = revenue.objects.filter(
                organization=self.target_organization,
                year=get_current_summary_report_year() - i,
                revenue_source__government_type="State").aggregate(sum=Sum('reported_value'))['sum']

            output_dict['Federal revenues'][year_x] = revenue.objects.filter(
                organization=self.target_organization,
                year=get_current_summary_report_year() - i,
                revenue_source__government_type="Federal").aggregate(sum=Sum('reported_value'))['sum']

            output_dict['Operating investments'][year_x] = transit_data.objects.filter(
                organization=self.target_organization,
                year=get_current_summary_report_year() - i,
                transit_metric_id=9).aggregate(sum=Sum('reported_value'))['sum']

            # output_dict['Operating investments'][year_x] = None_sum(output_dict['Operating investments'][year_x],
            #                                                         revenue.objects.filter(
            #                                                             organization=self.target_organization,
            #                                                             year=get_current_summary_report_year() - i,
            #                                                             revenue_source__funding_type="Operating").aggregate(sum=Sum('reported_value'))['sum'])

            output_dict['Local capital investments'][year_x] = expense.objects.filter(
                organization=self.target_organization,
                year=get_current_summary_report_year() - i,
                expense_source__heading='Local capital expenditures').aggregate(sum=Sum('reported_value'))['sum']

            output_dict['State capital investments'][year_x] = revenue.objects.filter(
                organization=self.target_organization,
                year=get_current_summary_report_year() - i,
                revenue_source__government_type="State",
                revenue_source__funding_type="Capital").aggregate(sum=Sum('reported_value'))['sum']

            output_dict['Federal capital investments'][year_x] = revenue.objects.filter(
                organization=self.target_organization,
                year=get_current_summary_report_year() - i,
                revenue_source__government_type="Federal",
                revenue_source__funding_type="Capital").aggregate(sum=Sum('reported_value'))['sum']

            output_dict['Other investments'][year_x] = expense.objects.filter(
                organization=self.target_organization,
                year=get_current_summary_report_year() - i,
                expense_source__id__in=[2, 4, 5]).aggregate(sum=Sum('reported_value'))['sum']

            i += 1
        revenue_table = []
        investment_table = []
        output = []
        for key, value in output_dict.items():
            row = [key, value['two_years_ago'], value['previous_year'], value['this_year']]
            print(row)
            if 'revenues' in key:
                revenue_table.append(row)
            else:
                investment_table.append(row)
        # print(revenue_table)
        revenue_table = [[v if v is not None else 0 for v in nested] for nested in revenue_table]
        # print('revenue_table:')
        # print(revenue_table)
        # # print(investment_table)
        investment_table = [[v if v is not None else 0 for v in nested] for nested in investment_table]

        revenue_table.append(self._calculate_total('Total revenues', revenue_table))
        revenue_table = self._calculate_percent(revenue_table)
        revenue_table = self._format_numbers(revenue_table)

        investment_table.append(self._calculate_total('Total investments', investment_table))
        investment_table = self._calculate_percent(investment_table)
        investment_table = self._format_numbers(investment_table)

        output.append(['Revenues'])
        output.extend(revenue_table)
        output.append(['Investments'])
        output.extend(investment_table)

        return output





class ExportReport_ReportType(SummaryBuilderReportType):
    """Class for holding each Report Type and its data"""

    def __init__(self, report_type):
        super().__init__(report_type)
        self.export_report_data_list = []
        self.pretty_report_type = self.get_pretty_report_type()

    def append_export_report_data(self, data):
        self.export_report_data_list.append(data)

####
# For anyone encountering this code in the future.  I'm deeply sorry. Sometimes the only way out is through.
# Maybe it will be refactored before you get read this.
#
# “Every now and then when your life gets complicated and the weasels start closing in, the only real cure is to load up on heinous chemicals (or write some shitty code) and then drive like a bastard from Hollywood to Las Vegas. To relax, as it were, in the womb of the desert sun.”
# # ― Hunter S. Thompson
####

class ExportReport_Data():
    def __init__(self, current_SummaryDataEntryBuilder):
        t, nav_headers = current_SummaryDataEntryBuilder.get_header_navigation()
        if isinstance(nav_headers, list):
            nav_headers = nav_headers[current_SummaryDataEntryBuilder.current_increment - 1]
            self.nav_headers_as_list = nav_headers
            self.nav_headers = ' '.join([str(elem) for elem in nav_headers])
            self.pretty_nav_headers = str(nav_headers[0]) + " (" + str(nav_headers[1]) + ")"
        else:
            self.nav_headers_as_list = [nav_headers, None]
            self.nav_headers = nav_headers
            self.pretty_nav_headers = nav_headers
        self.query_set = current_SummaryDataEntryBuilder.get_form_queryset().filter(
            year__gte=get_current_summary_report_year() - 3)
        self.current_builder = current_SummaryDataEntryBuilder
        self.year = get_current_summary_report_year()

    def _data_as_list_of_lists(self, include_comment=False, qry_set=None):
        """Reformats data to a tabular form where each list in the list is a row.  This is used to format data for
        easier printing, display, or manipulation"""
        col_data = {}
        list_of_lists = []
        i = 0
        if not qry_set:
            my_query_set = self.query_set
        for year_x in ['this_year', 'previous_year', 'two_years_ago']:
            col_data[year_x] = list(my_query_set.filter(year=self.year - i).values_list('reported_value', flat=True))
            i = i + 1

        if include_comment:
            comment = list(my_query_set.filter(year=self.year).values_list('comments', flat=True))
        col_labels = my_query_set.filter(year=self.year).values_list(
            self.current_builder.get_metric_model_name() + "__name", flat=True)
        for k, col in enumerate(col_labels):
            row = [col, col_data['two_years_ago'][k], col_data['previous_year'][k], col_data['this_year'][k]]
            if include_comment:
                if comment[k] is None:
                    comment[k] = ''
                row.extend([comment[k]])
            list_of_lists.append(row)

        return list_of_lists

    def _remove_empty_data(self, list_of_lists):
        for i in list_of_lists[:]:
            delete_row = False
            if all([x == 0 or x is None or x == 'None' for x in i[1:4]]):
                delete_row = True
            if delete_row:
                list_of_lists.remove(i)
        return list_of_lists

    def get_data(self, order_by_summary=False, remove_empty_data=False, none_to_zero=True, include_comment=True):
        if order_by_summary and self.current_builder.has_order_in_summary():
            self.query_set = self.query_set.order_by(
                self.current_builder.get_metric_model_name() + '__order_in_summary')
        output = self._data_as_list_of_lists(include_comment=include_comment)
        if remove_empty_data:
            output = self._remove_empty_data(output)
        if none_to_zero:
            output = [[v if v is not None else 0 for v in nested] for nested in output]

        return output

    def get_totals(self):
        # print(self.current_builder.report_type)
        if self.current_builder.report_type in ['transit_data']:
            self.can_total, self.totals_list = False, []
        else:
            totals = [0, 0, 0]
            j = 0
            for i in [2, 1, 0]:
                value = self.query_set.filter(year=self.year - i,
                                              reported_value__isnull=False).aggregate(Sum('reported_value'))['reported_value__sum']
                if value is None:
                    value = 0
                totals[j] = value
                j = j + 1
        return totals

    def get_farebox_revenue_as_list_of_lists(self):
        output = False
        if self.current_builder.report_type == "transit_data":
            qry_set = self.query_set.filter(transit_metrics_id=10)
            output = self._data_as_list_of_lists(qry_set)
        return output
